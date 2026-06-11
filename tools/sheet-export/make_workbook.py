#!/usr/bin/env python3
"""Bundle the CSVs exported by tools/sheet-export into one .xlsx workbook (one tab per CSV),
ready for Google Sheets: File -> Import -> Upload -> 'Replace spreadsheet'.

Usage: python3 make_workbook.py <csv-dir> <output.xlsx>
"""
import csv
import sys
from pathlib import Path

from openpyxl import Workbook

TAB_ORDER = ["LeagueDefinitions", "Quests", "Formations", "Legends"]  # small/tunable tabs first


def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__)
        return 1
    csv_dir, out_path = Path(sys.argv[1]), Path(sys.argv[2])

    wb = Workbook()
    wb.remove(wb.active)
    for tab in TAB_ORDER:
        path = csv_dir / f"{tab}.csv"
        ws = wb.create_sheet(title=tab)
        with path.open(newline="", encoding="utf-8") as f:
            for row in csv.reader(f):
                # Force every cell to string so Sheets keeps e.g. season "1995/96" verbatim.
                ws.append([str(cell) for cell in row])
        print(f"{tab}: {ws.max_row - 1} data rows")

    wb.save(out_path)
    print(f"Wrote {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
